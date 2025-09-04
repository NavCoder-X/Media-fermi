
# librerie
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import time
from sys import stdout
import sys, os
import platform
import subprocess
import os
from openpyxl import workbook 
from openpyxl.styles import PatternFill , Font ,Alignment ,Border ,Side

excel_path = "voti.xlsx"
numer_materie_path = "n_materie.txt"

# chek voti
def chek():
    
    # path
    def resource_path(relative_path):
        if hasattr(sys,"_MEIPASS"):
            return os.path.join(sys._MEIPASS,relative_path)
        return os.path.join(os.path.abspath("."),relative_path)

    driver_path = resource_path("chromedriver.exe")
    user_path = "login/id.txt"
    pass_path = "login/password.txt"
    browser_mode = resource_path("Browser_mode.txt")

    #controllo credenziali
    with open(user_path, "r") as file:
        codice = file.read().strip()
    if codice=="":
        print("nessun codice")
        return "nessun id"
    with open(pass_path, "r") as file:
        password = file.read().strip()
    if password=="":
        print("nessuna password trovata")
        return "nessuna password"
    
    # usa in background e altre opzioni
    option = Options()
    with open(browser_mode,"r") as f:
        status = f.read()
    if status=="0":
        option.add_argument("--headless")
    option.add_argument("--no-sandbox")
    option.add_argument("--disable-gpu")
    option.add_argument("--disable-dev-shm-usage")

    # setup driver
    service = Service(driver_path)  
    driver = webdriver.Chrome(service=service,options=option)


    # open the website
    driver.get("https://web.spaggiari.eu/home/app/default/login.php")

    # login
    user=driver.find_element(By.ID, "login")
    password_log=driver.find_element(By.ID, "password")
    user.send_keys(codice)
    password_log.send_keys(password)
    bottone = driver.find_element(By.CSS_SELECTOR, "button[type='submit']")
    bottone.click()

    # naviga fino al anno precedente
    time.sleep(10)
    try:
        bottone = driver.find_element(By.CSS_SELECTOR, "button[type='button']")
        bottone.click()
    except:
        pass
    try:
        anno_precedente = driver.find_element(By.CSS_SELECTOR, "p[class='voce_menu_colonna_sx']")
        anno_precedente.click()
    except:
        pass

    # Passa alla nuova finestra/scheda
    try:
        time.sleep(2)  # attesa breve per apertura finestra
        windows = driver.window_handles
        driver.switch_to.window(windows[-1])
    except:
        driver.quit()
        return "credenziali errate"
        
    time.sleep(10) # attesa per caricamento pagina

    # Trova tutte le righe della tabella con classe 'griglia rigtab'
    righe = driver.find_elements(By.CSS_SELECTOR, "tr[align='left']")
    if len(righe) >= 5:
        righe[4].click()  # Indici Python partono da 0, quindi la quinta ? [4]
    else:
        stdout.write("\033[91mCredenziali errate!\033[0m\n")
        driver.quit()
        return "credenziali errate"
    time.sleep(4)
    try:
    # Clicca sul bottone che contiene il testo "Tutto"
        tutto_button = driver.find_element(By.XPATH, "//button[contains(., 'Tutto')]")
        tutto_button.click()
    except:
        driver.quit()
        return "credenziali errate"

    time.sleep(8)

    # Trova tutti i voti nei paragrafi all'interno della struttura tr>td>div>p
    voti = []
    trs = driver.find_elements(By.TAG_NAME, "tr")
    for tr in trs:
        tcs = tr.find_elements(By.TAG_NAME, "td")
        for tc in tcs:
            if tc.get_attribute("colspan") == "48":
                voti.append(tr.text.strip())
            if tc.get_attribute("colspan") == "2":
                divs = tc.find_elements(By.TAG_NAME, "div")
                for div in divs:
                    if div.get_attribute("class").strip() == "cella_div cella_div f_reg_voto_dettaglio":
                        continue
                    paragrafi = div.find_elements(By.TAG_NAME, "p")
                    for p in paragrafi:
                        testo = p.text.strip()
                        if testo:  # solo paragrafi non vuoti
                            voti.append(testo)
    voti=voti[3:]
    driver.quit()

    # scrittura voti in excel file 
    materie_voto={}
    lista=[]
    iter=0

    # rendere tutti i voti float ed escludere religione
    for i in voti:
        i=i.strip()
        if i=="RELIGIONE CATTOLICA/ATTIVITA'ALTERNATIVA" or i=="O":
            continue
        if iter==0:
            lista.append(i)
            iter=1
        elif len(str(i))>5:
            try:
                voti_coronologici = lista[1:]
                voti_coronologici.reverse()  # Inverti l'ordine dei voti
                materie_voto[lista[0]]=voti_coronologici
            except:
                continue
            lista=[]
            lista.append(i)
        else:
            if "EDUCAZIONE CIVICA" in i:
                break
            if "-" in i:
                i=i.replace("-", "")
                i=float(i)-0.25
            elif "+" in i:
                i=i.replace("+", "")
                i=float(i)+0.25
            elif "½" in i:
                i=i.replace("½", "")
                i=float(i)+0.5
            else:
                try:
                    i=float(i)
                except:
                    continue
            lista.append(i)

    # scrittura di ed civica
    voti_coronologici = lista[1:]
    voti_coronologici.reverse()  # Inverti l'ordine dei voti
    materie_voto[lista[0]]=voti_coronologici

    # scrittura numero materie
    numero_materie = len(materie_voto.keys())
    with open(numer_materie_path,"w") as f:
        f.write(str(numero_materie))

   # creazione excel file
    wb = workbook.Workbook()
    ws = wb.active
    ws.title = "Voti"

    ws.append(['Materia', 'Voto_1','Voto_2','Voto_3','Voto_4','Voto_5','Voto_6','Voto_7','Voto_8','Voto_9','Voto_10', 'Media'])

    riga = 2
    for i in materie_voto.keys():
        voti=materie_voto[i]
        ws.append([i]+voti)
        ws[f"L{riga}"] = f"=AVERAGE(B{riga}:K{riga})"
        riga += 1

    # stile excel
    stile_allineamento = Alignment(horizontal="center",vertical="center")
    bold = Font(bold=True)
    fieldnames_background_color = PatternFill(start_color="99D6FB",fill_type="solid")
    materie_background_color = PatternFill(start_color="F6FE9F",fill_type="solid")
    voti_background_color = PatternFill(start_color="B6B6B6",fill_type="solid")
    medie_background_color = PatternFill(start_color="FFAFE1",fill_type="solid")
    thin = Side(style="thin",color="000000")
    bordo = Border(left=thin,top=thin,right=thin,bottom=thin)
    for row in ws.iter_rows(min_row=1,max_row=len(materie_voto.keys())+1,min_col=1,max_col=12):
        for cell in row:
            cell.alignment = stile_allineamento
            cell.border = bordo
            if cell.row==1:
                cell.fill = fieldnames_background_color
                cell.font = bold
            elif cell.row>1 and cell.column==1:
                cell.fill = materie_background_color
                cell.font = bold
            elif cell.row>1 and cell.column==12:
                cell.fill = medie_background_color
            else:
                cell.fill = voti_background_color
            
    ws.column_dimensions["A"].width = 55


    wb.save(excel_path)
    return "dati aggiornati"

def media():
    # calcolo media senza inculedere ed civica
    voti_processati = []
    flag = 0
    totale=0
    voti = get_data()

    for k in voti:
        for voto in k[0:len(k)-1]:
            if voto=="EDUCAZIONE CIVICA":
                flag=11
            elif flag>0:
                flag-1
                continue
            if len(str(voto))>5:
                continue
            elif voto==None:
                continue
            voti_processati.append(voto)
    n = len(voti_processati)

    for voto in voti_processati:
        try:
            voto = float(voto)
        except:
            print(f"Errore di conversione per il voto: {voto}")
            continue
        totale+=voto
        m=totale/n

    try:
        print("Media dei voti:", m)
        return m
    except:
        print("Nessun voto trovato o errore nel calcolo della media.")
        return "nessun voto trovato"
    
def quanto_posso_prendere():
    data = get_data()
    l = []

    for i in data:
        materia = i[0]
        media = round(i[-1],2)
        totale = 0
        esito = ""
        n = 1
        for j in i[1:-1]:
            if j == None:
                continue
            totale += float(j)
            n += 1
        target = n * 6
        v = target - totale
        if v < 6:
            esito = f"puoi prendere: {v:.2f}"
        else:
            esito = f"devi prendere: {v:.2f}"
        l.append(f"{materia} | media:{media} | {esito}")
    return l

def graficoXmateria(choice):
    data_x = []
    data_y = []
    n=0
    totale=0
    m=0
    flag=False
    voti = get_data()
    dati=[]

    for categoria in voti:
        for voto in categoria:
            try:
                voto = voto.strip()
            except:
                pass
            if choice==voto:
                flag=True
                continue
            if flag:
                if len(str(voto))>5 or voto==None:
                    break
                dati.append(voto)

    for voto in dati:
        n+=1
        totale+=voto
        m=totale/n
        data_x.append(n) # numero del voto
        data_y.append(m) # media progressiva
    return data_x , data_y 

def materie():

    data = []
    voti = get_data()
    for materia in voti:
        materia=materia[0]       
        data.append(materia)
    try:
        data.pop()
    except:
        data = "nessun dato"
    return data

def csv():
    if platform.system() == 'Windows':
        os.startfile(excel_path)
    elif platform.system() == 'Darwin':
        subprocess.call(['open', excel_path])
    else:
        subprocess.call(['xdg-open', excel_path])

def get_data():

    with open(numer_materie_path,"r") as f:
        numero_materie = int(f.read())

    lista_voti=[]
    voti=[]

    from openpyxl import load_workbook

    wb = load_workbook("voti.xlsx",data_only=True)
    ws = wb.active

    for row in ws.iter_rows(min_row=1,max_row=numero_materie,min_col=1,max_col=12):
        for cell in row:
            if cell.row==1:
                continue
            voti.append(cell.value)
            if cell.column==12:
                lista_voti.append(voti)
                voti=[]
    # calcolo manulae della media in quanto excel restituisce una formula
    for r in lista_voti:
        totale = 0
        n = 0
        for j in r[1:-1]:
            if j==None:
                continue
            totale+=j
            n+=1
        m=totale/n
        r[-1]=m
    return lista_voti


